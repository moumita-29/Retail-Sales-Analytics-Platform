"""
Retail Sales Analytics Platform
File: python/export_to_excel.py
Description: Export analytics data to a formatted Excel workbook
"""

import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Output path ──────────────────────────────────────────────
OUTPUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Retail_Sales_Report.xlsx")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Colors ───────────────────────────────────────────────────
ACCENT    = "4361EE"
HEADER_BG = "161B22"
DARK_BG   = "0D1117"
WHITE     = "C9D1D9"


# ── Styling helpers ──────────────────────────────────────────
def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill      = PatternFill("solid", fgColor=ACCENT)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = Border(bottom=Side(style="thin", color="30363D"))


def style_data(ws, end_row, num_cols):
    fill_a = PatternFill("solid", fgColor=HEADER_BG)
    fill_b = PatternFill("solid", fgColor=DARK_BG)
    for row in range(2, end_row + 1):
        fill = fill_a if row % 2 == 0 else fill_b
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill      = fill
            cell.font      = Font(color=WHITE, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def write_sheet(wb, name, df):
    ws = wb.create_sheet(title=name)
    ws.sheet_properties.tabColor = ACCENT

    for ci, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=ci, value=col_name)
    style_header(ws, len(df.columns))

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    style_data(ws, len(df) + 1, len(df.columns))
    auto_width(ws)
    ws.freeze_panes = "A2"


def create_kpi_sheet(wb, kpis):
    ws = wb.create_sheet(title="KPI Summary", index=0)
    ws.sheet_view.showGridLines = False

    bg = PatternFill("solid", fgColor=DARK_BG)
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=12):
        for cell in row:
            cell.fill = bg

    # Title
    ws.merge_cells("B2:J2")
    t = ws["B2"]
    t.value     = "RETAIL SALES ANALYTICS — KPI DASHBOARD"
    t.font      = Font(bold=True, size=18, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 40

    # KPI cards
    card_fill = PatternFill("solid", fgColor=HEADER_BG)
    items = list(kpis.items())
    for i, (label, value) in enumerate(items):
        col = 2 + i * 2
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        lc = ws.cell(row=5, column=col, value=label)
        lc.font = Font(bold=True, color=WHITE, size=10)
        lc.fill = card_fill
        lc.alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        vc = ws.cell(row=6, column=col, value=str(value))
        vc.font = Font(bold=True, color="4CC9F0", size=16)
        vc.fill = card_fill
        vc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[6].height = 32

    ws["B8"] = "Generated: " + pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
    ws["B8"].font = Font(italic=True, color="6E7681", size=9)


# ── Demo data generator ──────────────────────────────────────
def make_demo_data():
    np.random.seed(42)
    n       = 300
    cats    = ["Furniture", "Technology", "Office Supplies"]
    regions = ["North", "South", "East", "West", "Central"]
    segs    = ["Consumer", "Corporate", "Home Office"]

    cat     = np.random.choice(cats, n)
    revenue = np.random.uniform(100, 2000, n)
    cost_r  = np.where(cat == "Furniture", 0.45,
              np.where(cat == "Technology", 0.55, 0.35))
    cost    = (revenue * cost_r * np.random.uniform(0.8, 1.2, n)).round(2)
    profit  = (revenue - cost).round(2)
    revenue = revenue.round(2)

    dates   = pd.date_range("2024-01-01", periods=n, freq="D").to_numpy()
    sampled = np.random.choice(dates, n, replace=True)

    sales_df = pd.DataFrame({
        "order_id":         range(1, n + 1),
        "order_date":       pd.to_datetime(sampled),
        "customer_segment": np.random.choice(segs, n),
        "region_name":      np.random.choice(regions, n),
        "category":         cat,
        "quantity":         np.random.randint(1, 20, n),
        "discount_%":       np.random.choice([0, 5, 10, 15, 20], n),
        "revenue":          revenue,
        "cost":             cost,
        "profit":           profit,
        "profit_margin_%":  (profit / revenue * 100).round(2),
    })

    monthly_df = (
        sales_df.assign(month=sales_df["order_date"].dt.to_period("M").astype(str))
        .groupby("month")[["revenue", "profit"]]
        .sum().round(2).reset_index()
    )
    monthly_df["growth_%"] = (monthly_df["revenue"].pct_change() * 100).round(2)

    regional_df = (
        sales_df.groupby("region_name")
        .agg(total_orders=("order_id", "count"),
             total_revenue=("revenue", "sum"),
             total_profit=("profit", "sum"))
        .round(2).reset_index()
    )
    regional_df["margin_%"] = (regional_df["total_profit"] /
                                regional_df["total_revenue"] * 100).round(2)

    product_df = (
        sales_df.groupby("category")
        .agg(units_sold=("quantity", "sum"),
             total_revenue=("revenue", "sum"),
             total_profit=("profit", "sum"),
             avg_margin=("profit_margin_%", "mean"))
        .round(2).reset_index()
    )

    kpis = {
        "Total Revenue":    f"${sales_df['revenue'].sum():,.2f}",
        "Total Profit":     f"${sales_df['profit'].sum():,.2f}",
        "Profit Margin":    f"{sales_df['profit_margin_%'].mean():.1f}%",
        "Total Orders":     f"{sales_df['order_id'].nunique():,}",
        "Avg Order Value":  f"${sales_df['revenue'].mean():,.2f}",
    }
    return sales_df, monthly_df, regional_df, product_df, kpis


# ── Main ─────────────────────────────────────────────────────
if __name__ == "__main__":
    print("[→] Generating Excel report with demo data...")
    sales_df, monthly_df, regional_df, product_df, kpis = make_demo_data()

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    create_kpi_sheet(wb, kpis)
    write_sheet(wb, "Sales Data",      sales_df)
    write_sheet(wb, "Monthly Trends",  monthly_df)
    write_sheet(wb, "Regional Perf",   regional_df)
    write_sheet(wb, "Product Perf",    product_df)

    wb.save(OUTPUT_FILE)
    print(f"[✓] Excel report saved → {OUTPUT_FILE}")

    print("\n── KPI Summary ──────────────────────────────")
    for k, v in kpis.items():
        print(f"  {k:<20}: {v}")
    print("─────────────────────────────────────────────\n")
